# -*- coding: utf-8 -*-
"""검사성적서(Inspection Certificate) 엑셀 샘플 생성기.

목적: 데이터 모델 v2(입고검사 + 페인트후검사 2단계)를 기준으로 한 장(=로트 1건)짜리
      검사성적서 양식을 openpyxl 로 출력한다. 실제 값은 예시(샘플) 수치다.

설계 원칙(요구사항 매핑):
  - 검사항목은 '행(row)'으로 펼치는 롱포맷 → 항목이 늘어도 행만 추가, 양식 불변 (요구 4)
  - 입고검사/페인트후검사를 분리된 표로 두고, 페인트후검사는 미완료(점진 기록) 허용 표시
  - 수량 정합성 규칙을 참/거짓(boolean) 술어로 한 표에 노출 (로그인 없는 앱의 기록 신뢰성)
  - 검사기간(시작~종료), 공급사, 항목별 불량률 모두 표기 (요구 1·2·3·5)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── 공통 스타일 상수 ─────────────────────────────────────────
FONT_NAME = "맑은 고딕"  # 윈도우 한글 기본 폰트
THIN = Side(style="thin", color="808080")          # 일반 셀 테두리(회색)
MEDIUM = Side(style="medium", color="000000")       # 강조 테두리(검정)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_TOP_MEDIUM = Border(left=THIN, right=THIN, top=MEDIUM, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="2F5496")   # 진한 파랑(대제목)
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")  # 연한 파랑(섹션 제목)
FILL_HEADER = PatternFill("solid", fgColor="F2F2F2")   # 표 머리행(회색)
FILL_NA = PatternFill("solid", fgColor="D9D9D9")       # 해당없음(음영)
FILL_FALSE = PatternFill("solid", fgColor="FFC7CE")    # 정합성 거짓(연빨강)
FILL_KPI = PatternFill("solid", fgColor="FFF2CC")      # KPI 강조(연노랑)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center", wrap_text=True)


def style_cell(ws, coord, value="", *, bold=False, size=10, color="000000",
               align=ALIGN_C, fill=None, border=BORDER_THIN, italic=False):
    """셀 하나에 값 + 스타일을 한 번에 적용하는 헬퍼."""
    cell = ws[coord]
    cell.value = value
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)
    cell.alignment = align
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border
    return cell


def merge_box(ws, rng, value="", **kwargs):
    """병합 셀 + 스타일. 병합 영역 전체에 테두리를 둘러준다."""
    ws.merge_cells(rng)
    first = rng.split(":")[0]
    style_cell(ws, first, value, **kwargs)
    # 병합된 나머지 셀에도 테두리 적용(인쇄 시 테두리 끊김 방지)
    border = kwargs.get("border", BORDER_THIN)
    fill = kwargs.get("fill", None)
    for row in ws[rng]:
        for c in row:
            if border is not None:
                c.border = border
            if fill is not None:
                c.fill = fill


def section_title(ws, row, text):
    """A~F 전폭 섹션 제목 줄."""
    merge_box(ws, f"A{row}:F{row}", text, bold=True, size=11,
              align=ALIGN_L, fill=FILL_SECTION)
    ws.row_dimensions[row].height = 22


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "검사성적서"

    # 열 너비(6열 A~F) — A4 세로 1매 가정
    widths = {"A": 6, "B": 20, "C": 16, "D": 18, "E": 16, "F": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 페이지 설정: A4 세로, 폭 맞춤, 여백
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5

    r = 1  # 현재 행 포인터

    # ── 표제부(회사명 / 대제목 / 문서관리) ───────────────────
    merge_box(ws, f"A{r}:A{r+2}", "SecuGen\nKorea", bold=True, size=12,
              color="FFFFFF", fill=FILL_TITLE)
    merge_box(ws, f"B{r}:D{r+2}", "검사성적서\nINSPECTION CERTIFICATE", bold=True,
              size=16, color="FFFFFF", fill=FILL_TITLE)
    style_cell(ws, f"E{r}", "문서번호", bold=True, size=9, fill=FILL_HEADER, align=ALIGN_C)
    style_cell(ws, f"F{r}", "QC-IC-2026-0142", size=9, align=ALIGN_C)
    style_cell(ws, f"E{r+1}", "개정 / Rev.", bold=True, size=9, fill=FILL_HEADER, align=ALIGN_C)
    style_cell(ws, f"F{r+1}", "Rev. 0", size=9, align=ALIGN_C)
    style_cell(ws, f"E{r+2}", "페이지", bold=True, size=9, fill=FILL_HEADER, align=ALIGN_C)
    style_cell(ws, f"F{r+2}", "1 / 1", size=9, align=ALIGN_C)
    for i in range(3):
        ws.row_dimensions[r + i].height = 20
    r += 3

    # 문서상태(미완료 안내) 띠
    merge_box(ws, f"A{r}:F{r}",
              "문서상태: 잠정(DRAFT) — 페인트후검사 미완료(is_complete=false), 회수 완료 후 재발행 예정",
              bold=True, size=9, color="C00000", align=ALIGN_C, fill=FILL_HEADER)
    ws.row_dimensions[r].height = 18
    r += 1

    # ── 식별 정보(라벨|값|라벨|값 4열 그리드) ─────────────────
    section_title(ws, r, "■ 품목 · 공급사 · 입고 · 로트 정보")
    r += 1

    # (라벨, 값, 라벨, 값) 4칸 한 줄씩
    id_rows = [
        ("공급사 (Supplier)", "ALT (ACTIVE / 현행)", "프리즘종류", "PLASTIC (플라스틱 프리즘)"),
        ("공급사 LOT", "ALT-2406-A", "규격/적용모델", "FDU07 적용 / 12.0×8.0×6.0mm"),
        ("품목코드 (ERP)", "(미입력 — ERP 등록 후 기입)", "단위", "EA"),
        ("입고번호", "RCV-2026-0087", "입고일 (요구2)", "2026-06-10"),
        ("입고수량 (요구1)", "5,000 EA", "로트번호 ★발급단위", "L240610-03"),
        ("로트수량", "2,000 EA", "검사방식", "전수검사 (FULL) / AQL N/A"),
    ]
    for (l1, v1, l2, v2) in id_rows:
        style_cell(ws, f"A{r}", l1, bold=True, size=9, fill=FILL_HEADER, align=ALIGN_L)
        merge_box(ws, f"B{r}:C{r}", v1, size=9, align=ALIGN_L)
        style_cell(ws, f"D{r}", l2, bold=True, size=9, fill=FILL_HEADER, align=ALIGN_L)
        merge_box(ws, f"E{r}:F{r}", v2, size=9, align=ALIGN_L)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1  # 한 줄 여백

    # ── [블록 A] 입고검사 개요 ────────────────────────────────
    section_title(ws, r, "① 입고검사 (INCOMING / 투명 원자재 전수검사) — 검사 개요")
    r += 1
    a_cols = ["검사방식", "검사기간 (요구3)", "검사수량", "양품", "불량", "판정 · 검사원 · 완료"]
    for i, c in enumerate(a_cols):
        style_cell(ws, f"{get_column_letter(i+1)}{r}", c, bold=True, size=9, fill=FILL_HEADER)
    r += 1
    a_vals = ["전수 (FULL)", "2026-06-12 ~ 2026-06-13", "2,000 EA", "1,960 EA", "40 EA", "PASS · 홍길동 · 완료(✔)"]
    for i, v in enumerate(a_vals):
        style_cell(ws, f"{get_column_letter(i+1)}{r}", v, size=9)
    ws.row_dimensions[r].height = 18
    r += 1

    # ── [블록 A-1] 입고검사 항목별 불량(롱포맷) ───────────────
    merge_box(ws, f"A{r}:F{r}",
              "입고검사 항목별 불량 (검사항목 마스터를 행으로 — 항목 추가 시 행만 늘어남)",
              bold=True, size=9, align=ALIGN_L, fill=FILL_SECTION)
    r += 1
    d_cols = ["No", "검사항목", "분류", "불량수량", "불량률 (÷2,000)"]
    col_spans = [("A", "A"), ("B", "C"), ("D", "D"), ("E", "E"), ("F", "F")]
    for (c0, c1), name in zip(col_spans, d_cols):
        merge_box(ws, f"{c0}{r}:{c1}{r}", name, bold=True, size=9, fill=FILL_HEADER)
    r += 1
    incoming_items = [
        ("1", "원자재", "원자재", "22", "1.10%", False),
        ("2", "디그&찍힘", "외관", "12", "0.60%", False),
        ("3", "스크래치", "외관", "6", "0.30%", False),
        ("—", "페인트", "페인트", "해당없음", "/", True),
        ("—", "페인트 작업 불량", "도장", "해당없음", "/", True),
    ]
    for (no, item, cat, qty, rate, is_na) in incoming_items:
        fill = FILL_NA if is_na else None
        style_cell(ws, f"A{r}", no, size=9, fill=fill)
        merge_box(ws, f"B{r}:C{r}", item, size=9, align=ALIGN_L, fill=fill)
        style_cell(ws, f"D{r}", cat, size=9, fill=fill)
        style_cell(ws, f"E{r}", qty, size=9, fill=fill)
        style_cell(ws, f"F{r}", rate, size=9, fill=fill)
        ws.row_dimensions[r].height = 16
        r += 1
    # 합계행(굵은 상단 테두리) — 검산: ==defect_qty 40
    style_cell(ws, f"A{r}", "합계", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    merge_box(ws, f"B{r}:C{r}", "입고검사 불량 소계 [검산: == 검사불량합 40]", bold=True, size=9,
              align=ALIGN_L, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    style_cell(ws, f"D{r}", "—", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    style_cell(ws, f"E{r}", "40", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    style_cell(ws, f"F{r}", "2.00%", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    r += 2

    # ── [블록 B] 페인트 외주 수량 흐름 ────────────────────────
    section_title(ws, r, "② 페인트 외주 — 발송 → 회수 (입고검사 양품을 외주 도장)")
    r += 1
    b_cols = ["외주업체", "발송일", "발송수량", "회수일", "회수수량", "회수완료 · 비고"]
    for i, c in enumerate(b_cols):
        style_cell(ws, f"{get_column_letter(i+1)}{r}", c, bold=True, size=9, fill=FILL_HEADER)
    r += 1
    b_vals = ["(주)대성도장", "2026-06-14", "1,960 EA", "2026-06-18", "1,950 EA", "미완료(□) · 10EA 미회수"]
    for i, v in enumerate(b_vals):
        style_cell(ws, f"{get_column_letter(i+1)}{r}", v, size=9)
    ws.row_dimensions[r].height = 18
    r += 1
    # 정합성 검산 행
    style_cell(ws, f"A{r}", "[검산]", bold=True, size=8, fill=FILL_HEADER)
    merge_box(ws, f"B{r}:C{r}", "발송 ≤ 입고양품 : 참 (1,960 ≤ 1,960)", size=8, align=ALIGN_L, italic=True)
    merge_box(ws, f"D{r}:F{r}", "회수 ≤ 발송 : 참 (1,950 ≤ 1,960)", size=8, align=ALIGN_L, italic=True)
    r += 2

    # ── [블록 C] 페인트후검사 개요 ────────────────────────────
    section_title(ws, r, "③ 페인트후검사 (POST_PAINT / 전수검사) — 검사 개요  ※불량 나중에 점진 기록")
    r += 1
    for i, c in enumerate(a_cols):  # 입고검사와 동일 컬럼 구조 재사용
        style_cell(ws, f"{get_column_letter(i+1)}{r}", c.replace("양품", "양품(=최종)"), bold=True, size=9, fill=FILL_HEADER)
    r += 1
    c_vals = ["전수 (FULL)", "2026-06-19 ~ (진행중)", "1,950 EA", "1,930 EA", "20 EA", "CONDITIONAL · 김철수 · 미완료(□)"]
    for i, v in enumerate(c_vals):
        style_cell(ws, f"{get_column_letter(i+1)}{r}", v, size=9)
    ws.row_dimensions[r].height = 18
    r += 1

    # ── [블록 C-1] 페인트후검사 항목별 불량(롱포맷) ───────────
    merge_box(ws, f"A{r}:F{r}",
              "페인트후검사 항목별 불량 (디그&찍힘·스크래치는 양 단계 공통 → 단계별 분리 집계)",
              bold=True, size=9, align=ALIGN_L, fill=FILL_SECTION)
    r += 1
    d_cols2 = ["No", "검사항목", "분류", "불량수량", "불량률 (÷1,950)"]
    for (c0, c1), name in zip(col_spans, d_cols2):
        merge_box(ws, f"{c0}{r}:{c1}{r}", name, bold=True, size=9, fill=FILL_HEADER)
    r += 1
    postpaint_items = [
        ("4", "페인트", "페인트", "9", "0.46%", False),
        ("5", "페인트 작업 불량", "도장", "5", "0.26%", False),
        ("2", "디그&찍힘 (공통)", "외관", "4", "0.21%", False),
        ("3", "스크래치 (공통)", "외관", "2", "0.10%", False),
        ("—", "원자재", "원자재", "해당없음", "/", True),
    ]
    for (no, item, cat, qty, rate, is_na) in postpaint_items:
        fill = FILL_NA if is_na else None
        style_cell(ws, f"A{r}", no, size=9, fill=fill)
        merge_box(ws, f"B{r}:C{r}", item, size=9, align=ALIGN_L, fill=fill)
        style_cell(ws, f"D{r}", cat, size=9, fill=fill)
        style_cell(ws, f"E{r}", qty, size=9, fill=fill)
        style_cell(ws, f"F{r}", rate, size=9, fill=fill)
        ws.row_dimensions[r].height = 16
        r += 1
    style_cell(ws, f"A{r}", "합계(잠정)", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    merge_box(ws, f"B{r}:C{r}", "페인트후 불량 소계 [미완료=잠정]", bold=True, size=9,
              align=ALIGN_L, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    style_cell(ws, f"D{r}", "—", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    style_cell(ws, f"E{r}", "20", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    style_cell(ws, f"F{r}", "1.03%", bold=True, size=9, fill=FILL_HEADER, border=BORDER_TOP_MEDIUM)
    r += 2

    # ── [블록 D] 수량 정합성 검증(참/거짓 술어) ───────────────
    section_title(ws, r, "④ 수량 정합성 검증 (규칙을 참/거짓으로 — 기록 신뢰성)")
    r += 1
    merge_box(ws, f"A{r}:C{r}", "정합성 규칙 (술어)", bold=True, size=9, fill=FILL_HEADER, align=ALIGN_L)
    style_cell(ws, f"D{r}", "결과", bold=True, size=9, fill=FILL_HEADER)
    merge_box(ws, f"E{r}:F{r}", "실제값", bold=True, size=9, fill=FILL_HEADER)
    r += 1
    checks = [
        ("Σ로트수량 ≤ 입고수량", "참", "2,000 ≤ 5,000", False),
        ("(전수) 입고검사수량 == 로트수량", "참", "2,000 == 2,000", False),
        ("입고: 양품 + 불량 == 검사수량", "참", "1,960 + 40 == 2,000", False),
        ("입고: 불량 == Σ항목별불량", "참", "40 == 40", False),
        ("페인트 발송 ≤ 입고양품", "참", "1,960 ≤ 1,960", False),
        ("페인트 회수 ≤ 발송", "참", "1,950 ≤ 1,960", False),
        ("(전수) 페인트후검사수량 == 회수수량", "참", "1,950 == 1,950", False),
        ("페인트후: 양품+불량 ≤ 검사수량 (미완료=soft)", "참(잠정)", "1,930 + 20 ≤ 1,950", False),
    ]
    for (rule, result, actual, is_false) in checks:
        fill = FILL_FALSE if is_false else None
        merge_box(ws, f"A{r}:C{r}", rule, size=9, align=ALIGN_L, fill=fill)
        style_cell(ws, f"D{r}", result, bold=True, size=9, fill=fill,
                   color=("C00000" if is_false else "006100"))
        merge_box(ws, f"E{r}:F{r}", actual, size=9, fill=fill)
        ws.row_dimensions[r].height = 16
        r += 1
    r += 1

    # ── 종합 결과(KPI) ───────────────────────────────────────
    section_title(ws, r, "■ 종합 결과")
    r += 1
    summary = [
        ("입고검사 양품", "1,960 EA"),
        ("최종 양품 (페인트후)", "1,930 EA"),
        ("입고검사 수율 (÷로트)", "98.00% (1,960 / 2,000)"),
        ("전체 수율 (÷로트)", "96.50% (1,930 / 2,000)"),
        ("종합 불량률", "3.25% ((40+25) / 2,000)"),
        ("종합 판정", "조건부 합격 (CONDITIONAL) — 페인트후검사 진행중"),
    ]
    for i, (label, value) in enumerate(summary):
        emphasize = label in ("최종 양품 (페인트후)", "종합 판정")
        merge_box(ws, f"A{r}:B{r}", label, bold=True, size=9, fill=FILL_HEADER, align=ALIGN_L)
        merge_box(ws, f"C{r}:F{r}", value, bold=emphasize, size=10 if emphasize else 9,
                  align=ALIGN_L, fill=(FILL_KPI if emphasize else None))
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

    # ── 비고 ─────────────────────────────────────────────────
    section_title(ws, r, "■ 비고 및 검사근거")
    r += 1
    remarks = [
        ("비고", "페인트후검사 미완료(미회수 10EA) — 회수 시 추가 기록 후 재발행(Rev 증가) 예정."),
        ("검사근거/표준", "사내검사기준 QC-STD-PRISM Rev.2 / 도면 DWG-PR-0012"),
        ("발급 안내", "본 성적서는 Prism 검사관리 시스템에서 자동 생성됨 (데이터 기준 2026-06-24)"),
    ]
    for (label, value) in remarks:
        merge_box(ws, f"A{r}:B{r}", label, bold=True, size=9, fill=FILL_HEADER, align=ALIGN_L)
        merge_box(ws, f"C{r}:F{r}", value, size=9, align=ALIGN_L)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1

    # ── 결재란(검사 / 확인 / 승인) ───────────────────────────
    sign_labels = ["검사 (Inspected)\n검사원", "확인 (Checked)\n검사반장", "승인 (Approved)\n품질책임자"]
    spans = [("A", "B"), ("C", "D"), ("E", "F")]
    for (c0, c1), label in zip(spans, sign_labels):
        merge_box(ws, f"{c0}{r}:{c1}{r}", label, bold=True, size=9, fill=FILL_HEADER)
    r += 1
    for (c0, c1) in spans:  # 서명/날인 공란(높이 확보)
        merge_box(ws, f"{c0}{r}:{c1}{r}", "(서명 / 날인)", size=8, color="808080")
    ws.row_dimensions[r].height = 48
    r += 1

    # 인쇄 영역 지정
    ws.print_area = f"A1:F{r-1}"

    out = Path(__file__).resolve().parent / "검사성적서_샘플.xlsx"
    wb.save(out)
    print(f"[OK] 생성 완료: {out}")


if __name__ == "__main__":
    build()
