# -*- coding: utf-8 -*-
"""집계(report) 서비스 — 로트별/공급사별/항목별 집계 + 엑셀 다운로드.

요구 5: 입고일·로트별 양품/불량/검사항목별 불량률, 수율.
엑셀: 로트 1건 검사이력(건별 다운로드) + 전체 집계.
"""
from __future__ import annotations

import io

import db
import inspection_service


def _rate(defect, base):
    return round(defect / base * 100, 2) if base else 0.0


# ── 화면용 집계 ──────────────────────────────────────────────
def summary(data=None, product=None) -> dict:
    """전체 집계. product(item_code) 지정 시 해당 품목만 필터."""
    product = (product or "").strip() or None
    lot_filter = " AND pm.item_code = ?" if product else ""   # pm 조인이 있는 쿼리용
    lp = [product] if product else []
    conn = db.connect()
    try:
        # 로트별(입고검사 있는 로트)
        lots = conn.execute(
            "SELECT l.id, l.lot_no, l.lot_qty, r.receipt_no, r.receipt_date, "
            "       pm.item_code, pm.prism_type, pm.model, s.name AS supplier_name, "
            "       inc.start_date inc_s, inc.end_date inc_e, inc.inspected_qty inc_insp, inc.good_qty inc_good, inc.defect_qty inc_def, inc.is_complete inc_c, "
            "       post.good_qty post_good, post.defect_qty post_def, post.inspected_qty post_insp, post.is_complete post_c, "
            "       (SELECT COALESCE(SUM(sent_qty),0) FROM paint_job WHERE lot_id=l.id) sent, "
            "       (SELECT COALESCE(SUM(pr.returned_qty),0) FROM paint_return pr JOIN paint_job pj ON pj.id=pr.paint_job_id WHERE pj.lot_id=l.id) returned "
            "FROM lot l JOIN receipt r ON r.id=l.receipt_id JOIN prism_master pm ON pm.id=r.prism_id JOIN supplier s ON s.id=r.supplier_id "
            "LEFT JOIN inspection inc ON inc.lot_id=l.id AND inc.stage='INCOMING' "
            "LEFT JOIN inspection post ON post.lot_id=l.id AND post.stage='POST_PAINT' "
            "WHERE inc.id IS NOT NULL" + lot_filter + " ORDER BY r.receipt_date DESC, l.id DESC", lp).fetchall()
        by_lot = []
        for r in lots:
            d = dict(r)
            final_good = d["post_good"] if d["post_good"] is not None else d["inc_good"]
            by_lot.append({
                "lot_id": d["id"], "lot_no": d["lot_no"], "lot_qty": d["lot_qty"],
                "item_code": d["item_code"], "prism_type": d["prism_type"], "model": d["model"],
                "supplier_name": d["supplier_name"], "receipt_date": d["receipt_date"],
                "inc_period": f"{d['inc_s'] or ''} ~ {d['inc_e'] or '(진행중)'}",
                "inc_good": d["inc_good"], "inc_defect": d["inc_def"],
                "inc_rate": _rate(d["inc_def"] or 0, d["inc_insp"] or 0),
                "sent": d["sent"], "returned": d["returned"],
                "post_good": d["post_good"], "post_defect": d["post_def"],
                "post_rate": _rate(d["post_def"] or 0, d["post_insp"] or 0),
                "final_good": final_good,
                "yield": _rate(final_good or 0, d["lot_qty"]),
                "complete": bool(d["inc_c"]) and (d["post_c"] is None or bool(d["post_c"])),
            })

        # 로트 표기 간소화: 연월일_L## (그날 생성순으로 순번 → 같은 날 여러 로트도 유일)
        by_ymd = {}
        for b in by_lot:
            by_ymd.setdefault((b["receipt_date"] or "").replace("-", ""), []).append(b)
        for ymd, group in by_ymd.items():
            for seq, b in enumerate(sorted(group, key=lambda x: x["lot_id"]), start=1):
                b["lot_label"] = f"{ymd}_L{seq:02d}" if ymd else b["lot_no"]

        # 공급사별 — 입고합계와 검사합계를 따로 집계 후 병합(조인 중복합산 방지)
        recv = {r["supplier_id"]: r["s"] for r in conn.execute(
            "SELECT r.supplier_id, COALESCE(SUM(r.received_qty),0) s FROM receipt r "
            "JOIN prism_master pm ON pm.id=r.prism_id WHERE 1=1" + lot_filter + " GROUP BY r.supplier_id", lp).fetchall()}
        insp = {r["supplier_id"]: r for r in conn.execute(
            "SELECT r.supplier_id, COALESCE(SUM(inc.inspected_qty),0) inspected, "
            "       COALESCE(SUM(inc.good_qty),0) good, COALESCE(SUM(inc.defect_qty),0) defect "
            "FROM inspection inc JOIN lot l ON l.id=inc.lot_id JOIN receipt r ON r.id=l.receipt_id "
            "JOIN prism_master pm ON pm.id=r.prism_id "
            "WHERE inc.stage='INCOMING'" + lot_filter + " GROUP BY r.supplier_id", lp).fetchall()}
        by_supplier = []
        for s in conn.execute("SELECT id,name FROM supplier ORDER BY name").fetchall():
            i = insp.get(s["id"])
            inspected = i["inspected"] if i else 0
            good = i["good"] if i else 0
            defect = i["defect"] if i else 0
            if product and not recv.get(s["id"], 0) and not inspected:
                continue   # 선택 품목과 무관한 공급사는 제외
            by_supplier.append({"supplier_name": s["name"], "received": recv.get(s["id"], 0),
                                "inspected": inspected, "good": good, "defect": defect,
                                "defect_rate": _rate(defect, inspected)})

        # 항목별 불량률(단계 분리) — 품목 필터 시 lot→receipt→prism_master 조인으로 제한
        _insp_join = "JOIN lot l ON l.id=inc.lot_id JOIN receipt r ON r.id=l.receipt_id JOIN prism_master pm ON pm.id=r.prism_id "
        inc_total = conn.execute(
            "SELECT COALESCE(SUM(inc.inspected_qty),0) s FROM inspection inc " + _insp_join +
            "WHERE inc.stage='INCOMING'" + lot_filter, lp).fetchone()["s"]
        post_total = conn.execute(
            "SELECT COALESCE(SUM(inc.inspected_qty),0) s FROM inspection inc " + _insp_join +
            "WHERE inc.stage='POST_PAINT'" + lot_filter, lp).fetchone()["s"]
        # 도장완료 페인트후 배치(품목 레벨) — 선택 품목의 도장완료 프리즘만.
        # LEFT JOIN 팬아웃(도장완료 코드에 RAW 다수 매핑 시 이중합산) 방지: IN 서브쿼리로 필터.
        batch_where = (" AND pb.prism_id IN (SELECT painted_into_id FROM prism_master "
                       "WHERE item_code=? AND painted_into_id IS NOT NULL)") if product else ""
        bp = [product] if product else []
        post_total += conn.execute(
            "SELECT COALESCE(SUM(pb.good_qty+pb.defect_qty),0) s FROM paint_batch pb "
            "WHERE 1=1" + batch_where, bp).fetchone()["s"]
        items = conn.execute("SELECT id,name,category,applies_to_incoming,applies_to_post_paint "
                              "FROM inspection_item ORDER BY sort_order,id").fetchall()
        _def_join = "JOIN inspection i ON i.id=idf.inspection_id JOIN lot l ON l.id=i.lot_id JOIN receipt r ON r.id=l.receipt_id JOIN prism_master pm ON pm.id=r.prism_id "
        by_item = []
        for it in items:
            inc_def = conn.execute(
                "SELECT COALESCE(SUM(idf.defect_qty),0) s FROM inspection_defect idf " + _def_join +
                "WHERE i.stage='INCOMING' AND idf.inspection_item_id=?" + lot_filter, [it["id"]] + lp).fetchone()["s"]
            post_def = conn.execute(
                "SELECT COALESCE(SUM(idf.defect_qty),0) s FROM inspection_defect idf " + _def_join +
                "WHERE i.stage='POST_PAINT' AND idf.inspection_item_id=?" + lot_filter, [it["id"]] + lp).fetchone()["s"]
            post_def += conn.execute(
                "SELECT COALESCE(SUM(pbd.defect_qty),0) s FROM paint_batch_defect pbd "
                "JOIN paint_batch pb ON pb.id=pbd.batch_id "
                "WHERE pbd.inspection_item_id=?" + batch_where, [it["id"]] + bp).fetchone()["s"]
            by_item.append({
                "name": it["name"], "category": it["category"],
                "for_incoming": bool(it["applies_to_incoming"]), "for_post_paint": bool(it["applies_to_post_paint"]),
                "inc_defect": inc_def, "inc_rate": _rate(inc_def, inc_total),
                "post_defect": post_def, "post_rate": _rate(post_def, post_total),
            })

        # 품목 선택 목록(항상 전체 — 필터와 무관하게 드롭다운을 채움)
        products = [dict(r) for r in conn.execute(
            "SELECT pm.item_code item_code, MIN(pm.prism_type) prism_type, MIN(pm.model) model, "
            "       COUNT(DISTINCT l.id) lots "
            "FROM lot l JOIN receipt r ON r.id=l.receipt_id JOIN prism_master pm ON pm.id=r.prism_id "
            "JOIN inspection inc ON inc.lot_id=l.id AND inc.stage='INCOMING' "
            "WHERE pm.item_code IS NOT NULL AND pm.item_code <> '' "
            "GROUP BY pm.item_code ORDER BY pm.item_code").fetchall()]

        return {"ok": True, "product": product, "products": products,
                "by_lot": by_lot, "by_supplier": by_supplier, "by_item": by_item,
                "totals": {"received_total": sum(recv.values()), "inc_inspected": inc_total,
                           "post_inspected": post_total, "lot_count": len(by_lot)}}
    finally:
        conn.close()


# ── 엑셀 빌드 헬퍼 ───────────────────────────────────────────
def _wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _disp_len(v) -> int:
    """표시 폭(한글/CJK 는 2칸으로 계산)."""
    s = "" if v is None else str(v)
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in s)


def autofit(ws, min_w=6, max_w=60, pad=2):
    """열 너비를 내용 길이에 맞춤. 가로 병합 셀은 스팬 열에 나눠 반영(제목이 열을 부풀리지 않게)."""
    from openpyxl.utils import get_column_letter
    merges = list(ws.merged_cells.ranges)
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:      # 병합 비앵커/빈칸 제외
                continue
            lo, hi = c.column, c.column
            for m in merges:
                if m.min_row <= c.row <= m.max_row and m.min_col <= c.column <= m.max_col:
                    lo, hi = m.min_col, m.max_col
                    break
            share = _disp_len(c.value) / (hi - lo + 1)
            for col in range(lo, hi + 1):
                widths[col] = max(widths.get(col, 0), share)
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, w + pad))


def _style_header(ws, cells, fill="D9E1F2"):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in cells:
        ws[c].font = Font(name="맑은 고딕", bold=True, size=10)
        ws[c].fill = PatternFill("solid", fgColor=fill)
        ws[c].alignment = Alignment(horizontal="center", vertical="center")


def _simple_lot_label(lot_no):
    """RCV-YYYYMMDD-###-L## → YYYYMMDD_L## (없으면 원본)."""
    parts = (lot_no or "").split("-")
    ymd = next((p for p in parts if len(p) == 8 and p.isdigit()), None)
    lsuf = parts[-1] if parts and parts[-1].upper().startswith("L") else None
    return f"{ymd}_{lsuf}" if ymd and lsuf else (lot_no or "")


def lot_excel(lot_id, label=None):
    """로트 1건 검사이력 엑셀(건별 다운로드) — 테두리/음영/정렬 적용. (bytes, filename) 반환."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.cell.cell import MergedCell

    d = inspection_service.get_for_lot({"id": lot_id})
    if not d.get("ok"):
        return None, None
    L = d["lot"]
    disp = label or _simple_lot_label(L["lot_no"])   # 집계와 동일한 간단 라벨

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검사이력"
    ws.column_dimensions["A"].width = 16
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    FONT = "맑은 고딕"
    thin = Side(style="thin", color="000000")   # 검정 — 일반 '모든 테두리' 와 동일
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    FILL_SECT = PatternFill("solid", fgColor="D9E1F2")   # 섹션 제목
    FILL_LABEL = PatternFill("solid", fgColor="EEF2FA")   # 라벨
    FILL_HEAD = PatternFill("solid", fgColor="F3F6FC")   # 표 헤더
    C = Alignment(horizontal="center", vertical="center")
    LFT = RGT = C   # 전체 가운데 정렬로 통일(요청)
    MAXC = 6  # A..F

    def cell(coord, val=None, bold=False, size=10, fill=None, align=LFT):
        c = ws[coord]
        if not isinstance(c, MergedCell):   # 병합셀(앵커 외)은 value 읽기전용
            c.value = val
        c.font = Font(name=FONT, bold=bold, size=size)
        c.border = BORDER
        c.alignment = align
        if fill:
            c.fill = fill
        return c

    def fill_row(rownum, fill):
        # 병합 여백칸까지 테두리·음영을 채운다(값은 건드리지 않음)
        for col in "ABCDEF":
            c = ws[f"{col}{rownum}"]
            c.border = BORDER
            c.fill = fill

    r = 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAXC)
    cell("A1", f"프리즘 검사이력 — {disp}", bold=True, size=14, fill=FILL_SECT, align=C)
    fill_row(1, FILL_SECT)
    ws.row_dimensions[1].height = 26
    r = 3

    def kv(l1, v1, l2=None, v2=None):
        """A=라벨 B:C=값 / D=라벨 E:F=값 (테두리·병합)."""
        nonlocal r
        cell(f"A{r}", l1, bold=True, fill=FILL_LABEL, align=C)
        ws.merge_cells(f"B{r}:C{r}"); cell(f"B{r}", v1, align=LFT); cell(f"C{r}")
        if l2 is not None:
            cell(f"D{r}", l2, bold=True, fill=FILL_LABEL, align=C)
            ws.merge_cells(f"E{r}:F{r}"); cell(f"E{r}", v2, align=LFT); cell(f"F{r}")
        else:
            ws.merge_cells(f"D{r}:F{r}"); cell(f"D{r}"); cell(f"E{r}"); cell(f"F{r}")
        r += 1

    kv("로트번호", disp, "프리즘", f"{L.get('prism_type','')} {L.get('prism_spec','')}".strip())
    kv("공급사", L.get("supplier_name", ""), "로트수량", L.get("lot_qty", ""))
    kv("입고번호", L.get("receipt_no", ""))
    r += 1

    def sect(title):
        nonlocal r
        ws.merge_cells(f"A{r}:F{r}")
        cell(f"A{r}", title, bold=True, size=11, fill=FILL_SECT, align=LFT)
        fill_row(r, FILL_SECT)
        r += 1

    def insp_block(title, p):
        nonlocal r
        sect(title)
        kv("검사기간", f"{p.get('start_date','') or ''} ~ {p.get('end_date','') or '(진행중)'}", "검사수량", p.get("base_qty", 0))
        kv("양품", p.get("good_qty", 0), "불량", p.get("defect_qty", 0))
        # 항목 표 헤더 (A/B/C), D:F 병합 여백
        cell(f"A{r}", "검사항목", bold=True, fill=FILL_HEAD, align=C)
        cell(f"B{r}", "불량수량", bold=True, fill=FILL_HEAD, align=C)
        cell(f"C{r}", "불량률(%)", bold=True, fill=FILL_HEAD, align=C)
        ws.merge_cells(f"D{r}:F{r}"); cell(f"D{r}", fill=FILL_HEAD); cell(f"E{r}", fill=FILL_HEAD); cell(f"F{r}", fill=FILL_HEAD)
        r += 1
        base = p.get("base_qty", 0) or 0
        items = p.get("items", [])
        if not items:
            ws.merge_cells(f"A{r}:F{r}"); cell(f"A{r}", "(검사 항목 없음)", align=C); [cell(f"{c}{r}") for c in "BCDEF"]; r += 1
        for it in items:
            q = p.get("defects", {}).get(it["id"], 0)
            cell(f"A{r}", it["name"], align=LFT)
            cell(f"B{r}", q, align=RGT)
            cell(f"C{r}", _rate(q, base), align=RGT)
            ws.merge_cells(f"D{r}:F{r}"); cell(f"D{r}"); cell(f"E{r}"); cell(f"F{r}")
            r += 1
        r += 1

    insp_block("① 입고검사", d["incoming"])
    if d.get("is_plastic"):
        sect("② 페인트 외주")
        kv("발송합", d["paint"]["sent_sum"], "회수합", d["paint"]["returned_sum"])
        r += 1
        if d.get("post_paint") and d["post_paint"].get("exists"):
            insp_block("③ 페인트후검사", d["post_paint"])

    sect("종합")
    kv("최종 양품", d.get("final_good"), "수율(%)", _rate(d.get("final_good") or 0, L.get("lot_qty", 0)))

    autofit(ws)   # 텍스트 길이에 맞춰 열 너비 자동
    return _wb_bytes(wb), f"검사이력_{disp}.xlsx"


def report_excel(data=None, product=None):
    """전체 집계 엑셀 — 테두리 + 색구분(양품 초록/불량 빨강) + 입고일 기준(로트번호 제외).
    product(item_code) 지정 시 해당 품목만."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    s = summary(product=product)
    tag = f"_{s['product']}" if s.get("product") else ""
    wb = openpyxl.Workbook()

    FONT = "맑은 고딕"
    thin = Side(style="thin", color="000000")   # 검정 — 일반 '모든 테두리' 와 동일
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    F_TITLE = PatternFill("solid", fgColor="305496")   # 진파랑(제목)
    F_HEAD = PatternFill("solid", fgColor="4472C4")    # 파랑(헤더)
    F_GOOD = PatternFill("solid", fgColor="E2EFDA")    # 연초록(양품)
    F_BAD = PatternFill("solid", fgColor="FCE4D6")     # 연주황(불량)
    C = Alignment(horizontal="center", vertical="center")
    RGT = Alignment(horizontal="right", vertical="center")
    LFT = Alignment(horizontal="left", vertical="center")

    def styled_sheet(ws, title, headers, rows, good_idx=(), bad_idx=(), num_idx=()):
        ncol = len(headers)
        # 제목행(병합)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        t = ws.cell(1, 1, title)
        t.font = Font(name=FONT, bold=True, size=13, color="FFFFFF"); t.fill = F_TITLE; t.alignment = C
        for c in range(1, ncol + 1):
            cc = ws.cell(1, c); cc.fill = F_TITLE; cc.border = BORDER
        ws.row_dimensions[1].height = 22
        # 헤더행
        for c, h in enumerate(headers, 1):
            cc = ws.cell(2, c, h)
            cc.font = Font(name=FONT, bold=True, color="FFFFFF"); cc.fill = F_HEAD; cc.alignment = C; cc.border = BORDER
        # 데이터행
        for ri, row in enumerate(rows, 3):
            for ci, val in enumerate(row, 1):
                cc = ws.cell(ri, ci, val)
                cc.font = Font(name=FONT); cc.border = BORDER
                cc.alignment = C   # 전체 가운데 정렬
                if (ci - 1) in good_idx:
                    cc.fill = F_GOOD
                elif (ci - 1) in bad_idx:
                    cc.fill = F_BAD

    # ① 입고일별(로트번호 제외)
    ws1 = wb.active
    ws1.title = "입고일별 집계"
    h1 = ["입고일", "프리즘", "공급사", "입고수량", "입고양품", "입고불량", "입고불량률(%)", "최종양품", "수율(%)"]
    rows1 = [[b["receipt_date"], f"{b['prism_type']} {b.get('model') or ''}".strip(), b["supplier_name"],
              b["lot_qty"], b["inc_good"], b["inc_defect"], b["inc_rate"], b["final_good"], b["yield"]]
             for b in s["by_lot"]]
    t1 = "입고일별 검사 집계" + (f" — 품목 {s['product']}" if tag else "")
    styled_sheet(ws1, t1, h1, rows1, good_idx={4, 7}, bad_idx={5, 6}, num_idx={3, 4, 5, 6, 7, 8})
    autofit(ws1)

    # ② 공급사별
    ws2 = wb.create_sheet("공급사별")
    h2 = ["공급사", "입고수량", "검사수량", "양품", "불량", "불량률(%)"]
    rows2 = [[b["supplier_name"], b["received"], b["inspected"], b["good"], b["defect"], b["defect_rate"]]
             for b in s["by_supplier"]]
    styled_sheet(ws2, "공급사별 집계", h2, rows2, good_idx={3}, bad_idx={4}, num_idx={1, 2, 3, 4, 5})
    autofit(ws2)

    # ③ 입고검사 불량 (투명·원자재 프리즘) — 입고 적용 항목만
    ws3 = wb.create_sheet("입고검사불량")
    h3 = ["검사항목", "분류", "입고불량", "입고불량률(%)"]
    rows3 = [[b["name"], b["category"], b["inc_defect"], b["inc_rate"]]
             for b in s["by_item"] if b.get("for_incoming")]
    styled_sheet(ws3, "입고검사 불량 (투명·원자재 프리즘)", h3, rows3, bad_idx={2}, num_idx={2, 3})
    autofit(ws3)

    # ④ 페인트후 불량 (도장완료 프리즘) — 페인트후 적용 항목만
    ws4 = wb.create_sheet("페인트후불량")
    h4 = ["검사항목", "분류", "페인트후불량", "페인트후불량률(%)"]
    rows4 = [[b["name"], b["category"], b["post_defect"], b["post_rate"]]
             for b in s["by_item"] if b.get("for_post_paint")]
    styled_sheet(ws4, "페인트후 불량 (도장완료 프리즘)", h4, rows4, bad_idx={2}, num_idx={2, 3})
    autofit(ws4)

    return _wb_bytes(wb), f"프리즘_집계{tag}.xlsx"
