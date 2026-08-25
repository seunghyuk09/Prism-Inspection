# -*- coding: utf-8 -*-
"""ERP 재고수불부 임포트 — 소모(출고) → 소비 등록.

엑셀 포맷(ECount 재고수불부): A=일자, B=거래처명, C=적요, D=입고수량, E=출고수량, F=재고수량.
'소모' 거래처(소모II) 줄의 출고수량을 선택한 도장완료 프리즘의 소비로 등록한다.
미리보기(preview) → 확정(commit). 이미 등록된(같은 일자·수량) 소비는 자동 건너뜀(중복방지).
"""
from __future__ import annotations

import base64
import io

import db
import rules

COL_DATE, COL_VENDOR, COL_DESC, COL_IN, COL_OUT = 0, 1, 2, 3, 4
TAG = "[수불부임포트]"


def _load_wb(b64: str):
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl 라이브러리가 없습니다."
    try:
        raw = base64.b64decode((b64 or "").split(",")[-1])
        # read_only=False: 일부 ERP export 는 dimension 메타가 부실해 read_only 에서 1행만 읽힘
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True), None
    except Exception as exc:  # noqa: BLE001
        return None, f"엑셀을 읽지 못했습니다: {exc}"


def _fmt_date(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if " " in s:
        s = s.split(" ")[0]
    return s.replace("/", "-")[:10]


def _parse(wb, sheet) -> list[dict]:
    ws = wb[sheet]
    lines = []
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) <= COL_OUT:
            continue
        vendor = str(row[COL_VENDOR] or "")
        out = rules._to_int(row[COL_OUT])
        if out <= 0 or "소모" not in vendor:   # 소모II 라인만(계/합계/생산 제외)
            continue
        lines.append({"date": _fmt_date(row[COL_DATE]),
                      "product": str(row[COL_DESC] or "").strip(), "qty": out})
    return lines


def _painted_default(conn):
    r = conn.execute("SELECT id FROM prism_master WHERE paint_state='PAINTED' AND item_code LIKE '%ALT%' "
                     "ORDER BY id LIMIT 1").fetchone()
    if not r:
        r = conn.execute("SELECT id FROM prism_master WHERE paint_state='PAINTED' ORDER BY id LIMIT 1").fetchone()
    return r["id"] if r else None


def _sig(date, qty, product=None) -> tuple:
    # 중복판정은 (일자, 수량) 기준. 제품명은 표기 차이(예: 'w/Type' vs 'w/Type C Cable')로
    # 같은 소비를 재등록해버릴 수 있어 시그니처에 넣지 않는다(같은 소비 재등록 방지가 우선).
    return (str(date), int(qty))


def _existing_sigs(conn, prism_id):
    """이미 등록된 소비의 (일자, 수량) 집합 — 중복판정용(같은 소비 재등록 방지)."""
    sigs = set()
    for r in conn.execute("SELECT consumed_at, qty FROM consumption WHERE prism_id=?", (prism_id,)).fetchall():
        sigs.add(_sig(r["consumed_at"], r["qty"]))
    return sigs


def preview(data: dict) -> dict:
    wb, err = _load_wb(data.get("content"))
    if err:
        return {"ok": False, "error": err}
    sheet = data.get("sheet") or (wb.sheetnames[0] if wb.sheetnames else None)
    if sheet not in wb.sheetnames:
        return {"ok": False, "error": f"시트를 찾을 수 없습니다: {sheet}"}
    conn = db.connect()
    try:
        prism_id = data.get("prism_id") or _painted_default(conn)
        sigs = _existing_sigs(conn, prism_id) if prism_id else set()
        lines = _parse(wb, sheet)
        for l in lines:
            l["status"] = "DUP" if _sig(l["date"], l["qty"], l["product"]) in sigs else "NEW"
        prisms = [dict(r) for r in conn.execute(
            "SELECT id,item_code,spec FROM prism_master WHERE is_active=1 AND paint_state='PAINTED' ORDER BY item_code").fetchall()]
    finally:
        conn.close()
    new = [l for l in lines if l["status"] == "NEW"]
    return {"ok": True, "sheets": wb.sheetnames, "sheet": sheet, "prism_id": prism_id,
            "prisms": prisms, "lines": lines,
            "summary": {"total": len(lines), "new": len(new), "dup": len(lines) - len(new),
                        "consumed": sum(l["qty"] for l in new)}}


def commit(data: dict) -> dict:
    wb, err = _load_wb(data.get("content"))
    if err:
        return {"ok": False, "error": err}
    sheet = data.get("sheet") or (wb.sheetnames[0] if wb.sheetnames else None)
    if sheet not in wb.sheetnames:
        return {"ok": False, "error": "시트를 선택하세요."}
    conn = db.connect()
    try:
        prism_id = data.get("prism_id") or _painted_default(conn)
        if not prism_id:
            return {"ok": False, "error": "도장완료 프리즘을 선택하세요."}
        sigs = _existing_sigs(conn, prism_id)
        ts = db.now_str()
        made = 0
        skipped = 0
        total = 0
        for l in _parse(wb, sheet):
            sig = _sig(l["date"], l["qty"], l["product"])
            if sig in sigs:
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO consumption(prism_id,supplier_id,source,qty,consumed_at,note,created_at) "
                "VALUES (?,?, 'MANUAL', ?, ?, ?, ?)",
                (prism_id, None, l["qty"], l["date"], f"도장완료 소비 — {l['product']} {TAG}", ts))
            sigs.add(sig)   # 같은 파일 내 완전중복(일자·수량·제품 동일)만 방지
            made += 1
            total += l["qty"]
        conn.commit()
        return {"ok": True, "created": made, "skipped": skipped, "consumed": total}
    finally:
        conn.close()
