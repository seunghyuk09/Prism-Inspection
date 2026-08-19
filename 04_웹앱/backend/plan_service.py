# -*- coding: utf-8 -*-
"""구매계획서(production plan) 임포트 — 소비(프리즘 차감) 생성.

엑셀 포맷: 월별 시트, 데이터 7행~, C열=PRODUCT(제품코드), E열=Q'ty(생산수량).
제품마스터로 매핑 → 소비량 = 생산수량 × 대당 소요량 → 제품의 prism_id(보통 도장완료/유리) 차감.
미리보기(preview) 후 확정(commit). 확정은 서버가 b64 를 재파싱해 신뢰값으로 기록.
xlsx 파싱은 openpyxl(설치돼 있음) 사용.
"""
from __future__ import annotations

import base64
import io

import db
import rules

COL_PRODUCT = 2   # C열(0-based)
COL_QTY = 4       # E열(0-based)
DATA_START_ROW = 7


def _load_wb(b64: str):
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl 라이브러리가 없습니다(엑셀 파싱 불가)."
    try:
        raw = base64.b64decode((b64 or "").split(",")[-1])   # data URL 접두 제거
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        return wb, None
    except Exception as exc:  # noqa: BLE001
        return None, f"엑셀을 읽지 못했습니다: {exc}"


def _product_map(conn) -> dict:
    rows = conn.execute(
        "SELECT p.id, p.product_code, p.product_name, p.prism_id, p.prism_per_unit, "
        "       pm.item_code AS prism_code, pm.spec AS prism_spec "
        "FROM product p LEFT JOIN prism_master pm ON pm.id=p.prism_id WHERE p.is_active=1").fetchall()
    return {(r["product_code"] or "").strip(): dict(r) for r in rows}


def _parse(wb, sheet, conn) -> dict:
    ws = wb[sheet]
    pmap = _product_map(conn)
    lines = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or len(row) <= COL_QTY:
            continue
        code = row[COL_PRODUCT]
        if code is None or not str(code).strip():
            continue
        code = str(code).strip()
        qty = rules._to_int(row[COL_QTY])
        if qty <= 0:
            continue
        prod = pmap.get(code)
        if not prod:
            lines.append({"product_code": code, "qty": qty, "status": "UNMATCHED",
                          "prism_code": None, "consumed": 0})
        elif not prod["prism_id"]:
            lines.append({"product_code": code, "qty": qty, "status": "NO_PRISM",
                          "prism_code": None, "consumed": 0, "product_id": prod["id"]})
        else:
            consumed = rules.plan_consumption_qty(qty, prod["prism_per_unit"], prod["prism_id"])
            lines.append({"product_code": code, "qty": qty, "status": "OK",
                          "prism_code": prod["prism_code"], "prism_id": prod["prism_id"],
                          "consumed": consumed, "product_id": prod["id"]})
    matched = [l for l in lines if l["status"] == "OK"]
    summary = {
        "total_lines": len(lines), "matched": len(matched),
        "unmatched": len([l for l in lines if l["status"] == "UNMATCHED"]),
        "no_prism": len([l for l in lines if l["status"] == "NO_PRISM"]),
        "total_consumed": sum(l["consumed"] for l in matched),
    }
    return {"lines": lines, "summary": summary}


def preview(data: dict) -> dict:
    wb, err = _load_wb(data.get("content"))
    if err:
        return {"ok": False, "error": err}
    sheets = list(wb.sheetnames)
    sheet = data.get("sheet")
    if not sheet:
        cand = [s for s in sheets if s.lower() != "sheet1"]
        sheet = cand[-1] if cand else (sheets[-1] if sheets else None)
    if sheet not in sheets:
        return {"ok": False, "error": f"시트를 찾을 수 없습니다: {sheet}"}
    conn = db.connect()
    try:
        parsed = _parse(wb, sheet, conn)
    finally:
        conn.close()
    return {"ok": True, "sheets": sheets, "sheet": sheet, **parsed}


def commit(data: dict) -> dict:
    """확정 — 서버가 b64 를 재파싱해 production_plan + lines + consumption 기록."""
    wb, err = _load_wb(data.get("content"))
    if err:
        return {"ok": False, "error": err}
    sheet = data.get("sheet")
    if not sheet or sheet not in wb.sheetnames:
        return {"ok": False, "error": "시트를 선택하세요."}
    plan_month = (data.get("plan_month") or sheet).strip()
    is_final = 1 if data.get("is_final") else 0
    filename = (data.get("filename") or "").strip()

    conn = db.connect()
    try:
        parsed = _parse(wb, sheet, conn)
        ts = db.now_str()
        cur = conn.execute(
            "INSERT INTO production_plan(plan_month,source_file,is_final,imported_at,note,created_at,updated_at) "
            "VALUES (?,?,?,?,?,?,?)", (plan_month, filename, is_final, ts, (data.get("note") or "").strip(), ts, ts))
        plan_id = cur.lastrowid
        consumed_total = 0
        for l in parsed["lines"]:
            conn.execute(
                "INSERT INTO production_plan_line(plan_id,product_code,planned_qty,product_id,prism_id,consumed_qty,note) "
                "VALUES (?,?,?,?,?,?,?)",
                (plan_id, l["product_code"], l["qty"], l.get("product_id"), l.get("prism_id"),
                 l.get("consumed", 0), l["status"]))
            if l["status"] == "OK" and l.get("prism_id") and l.get("consumed", 0) > 0:
                conn.execute(
                    "INSERT INTO consumption(prism_id,supplier_id,source,source_plan_id,qty,consumed_at,note,created_at) "
                    "VALUES (?,?, 'PLAN', ?, ?, ?, ?, ?)",
                    (l["prism_id"], None, plan_id, l["consumed"], plan_month, l["product_code"], ts))
                consumed_total += l["consumed"]
        conn.commit()
        return {"ok": True, "plan_id": plan_id, "consumed_total": consumed_total, "summary": parsed["summary"]}
    finally:
        conn.close()


def list_plans(data=None) -> list[dict]:
    conn = db.connect()
    try:
        rows = conn.execute(
            "SELECT pp.id, pp.plan_month, pp.source_file, pp.is_final, pp.imported_at, pp.note, "
            "  (SELECT COALESCE(SUM(qty),0) FROM consumption WHERE source_plan_id=pp.id) AS consumed_total, "
            "  (SELECT COUNT(*) FROM production_plan_line WHERE plan_id=pp.id) AS line_count "
            "FROM production_plan pp ORDER BY pp.id DESC").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def delete_plan(data: dict) -> dict:
    """계획 취소 — 연결된 소비(차감)도 함께 되돌린다."""
    plan_id = data.get("id")
    conn = db.connect()
    try:
        conn.execute("DELETE FROM consumption WHERE source_plan_id=?", (plan_id,))
        conn.execute("DELETE FROM production_plan_line WHERE plan_id=?", (plan_id,))
        cur = conn.execute("DELETE FROM production_plan WHERE id=?", (plan_id,))
        conn.commit()
        if cur.rowcount == 0:
            return {"ok": False, "error": "계획을 찾을 수 없습니다."}
        return {"ok": True}
    finally:
        conn.close()
