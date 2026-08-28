# -*- coding: utf-8 -*-
"""도장완료 페인트후 불량(품목 레벨 일괄) — 우경 반납일별 집계.

로트 단위가 아닌 '도장완료 프리즘' 품목 레벨로 반납일별 양품/불량(항목별)을 기록.
집계(report)의 '페인트후 불량'에 반영된다(report_service 가 paint_batch* 를 합산).
입력: 수동(항목별 수량) 또는 우경 불량 엑셀 업로드(모든 시트).
"""
from __future__ import annotations

import base64
import io

import db
import rules

# 우경 자료 유형명 → 시스템 검사항목 별칭
_ALIAS = {"dig": "디그&찍힘"}


def _norm(s: str) -> str:
    return (s or "").strip().replace(" ", "").lower()


def painted_prisms(data=None) -> list[dict]:
    """도장완료(PAINTED) 프리즘 목록 — 선택용."""
    conn = db.connect()
    try:
        rows = conn.execute(
            "SELECT id,item_code,prism_type,model,spec FROM prism_master "
            "WHERE is_active=1 AND paint_state='PAINTED' ORDER BY item_code").fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def _find_or_create_item(conn, raw_name: str, ts: str) -> int:
    """유형명을 페인트후 검사항목에 매핑(없으면 생성). 페인트후 적용 보장."""
    name = (raw_name or "").strip()
    key = _norm(name)
    row = None
    target = _ALIAS.get(key)
    if target:
        row = conn.execute("SELECT id,applies_to_post_paint FROM inspection_item WHERE name=?", (target,)).fetchone()
    if not row:
        for it in conn.execute("SELECT id,name,applies_to_post_paint FROM inspection_item").fetchall():
            if _norm(it["name"]) == key:
                row = it
                break
    if row:
        if not row["applies_to_post_paint"]:
            conn.execute("UPDATE inspection_item SET applies_to_post_paint=1,updated_at=? WHERE id=?", (ts, row["id"]))
        return row["id"]
    order = conn.execute("SELECT COALESCE(MAX(sort_order),0)+1 AS n FROM inspection_item").fetchone()["n"]
    cur = conn.execute(
        "INSERT INTO inspection_item(name,category,applies_to_incoming,applies_to_post_paint,sort_order,is_active,created_at,updated_at) "
        "VALUES (?,?,0,1,?,1,?,?)", (name, "도장", order, ts, ts))
    return cur.lastrowid


def list_batches(data=None) -> list[dict]:
    prism_id = (data or {}).get("prism_id") if isinstance(data, dict) else None
    conn = db.connect()
    try:
        where = "WHERE pb.prism_id=?" if prism_id else ""
        params = (prism_id,) if prism_id else ()
        rows = conn.execute(
            "SELECT pb.id,pb.prism_id,pb.batch_date,pb.good_qty,pb.defect_qty,pb.note,pb.created_at, "
            "       pm.item_code, pm.spec "
            "FROM paint_batch pb JOIN prism_master pm ON pm.id=pb.prism_id " + where +
            " ORDER BY pb.batch_date DESC, pb.id DESC", params).fetchall()
        out = []
        for r in rows:
            defs = conn.execute(
                "SELECT ii.id AS item_id, ii.name, pbd.defect_qty FROM paint_batch_defect pbd "
                "JOIN inspection_item ii ON ii.id=pbd.inspection_item_id WHERE pbd.batch_id=? ORDER BY pbd.id",
                (r["id"],)).fetchall()
            d = dict(r)
            d["total"] = (r["good_qty"] or 0) + (r["defect_qty"] or 0)
            d["defects"] = [dict(x) for x in defs]
            out.append(d)
        return out
    finally:
        conn.close()


def save(data: dict) -> dict:
    """수동 저장 — {prism_id, batch_date, good_qty, defects:[{item_id,qty}], note}."""
    prism_id = data.get("prism_id")
    batch_date = (data.get("batch_date") or "").strip()
    good_qty = rules._to_int(data.get("good_qty"))
    note = (data.get("note") or "").strip()
    if not prism_id or not batch_date:
        return {"ok": False, "error": "도장완료 프리즘과 반납일을 입력하세요."}
    defects = []
    dq = 0
    for d in (data.get("defects") or []):
        q = rules._to_int(d.get("qty"))
        iid = d.get("item_id")
        if q <= 0 or not iid:
            continue
        defects.append((int(iid), q))
        dq += q
    conn = db.connect()
    try:
        dup = conn.execute("SELECT id FROM paint_batch WHERE prism_id=? AND batch_date=?",
                           (prism_id, batch_date)).fetchone()
        if dup:
            return {"ok": False, "error": f"이미 {batch_date} 배치가 있습니다. 삭제 후 다시 등록하세요."}
        ts = db.now_str()
        cur = conn.execute(
            "INSERT INTO paint_batch(prism_id,batch_date,good_qty,defect_qty,note,created_at) VALUES (?,?,?,?,?,?)",
            (prism_id, batch_date, good_qty, dq, note, ts))
        bid = cur.lastrowid
        for iid, q in defects:
            conn.execute("INSERT INTO paint_batch_defect(batch_id,inspection_item_id,defect_qty) VALUES (?,?,?)",
                         (bid, iid, q))
        conn.commit()
        return {"ok": True, "batch_id": bid, "good_qty": good_qty, "defect_qty": dq}
    finally:
        conn.close()


def update(data: dict) -> dict:
    """반납일 배치 수량만 그 자리에서 수정 — {id, good_qty, defects:[{item_id,qty}], note?}.

    삭제·재등록 없이 양품/불량유형별 수량을 고친다. 불량유형 목록은 통째로 교체
    (qty 0 은 제외). 집계(페인트후 불량)에 즉시 반영된다.
    """
    bid = data.get("id")
    if not bid:
        return {"ok": False, "error": "배치 ID가 없습니다."}
    good_qty = rules._to_int(data.get("good_qty"))
    defects = []
    dq = 0
    for d in (data.get("defects") or []):
        q = rules._to_int(d.get("qty"))
        iid = d.get("item_id")
        if q <= 0 or not iid:
            continue
        defects.append((int(iid), q))
        dq += q
    conn = db.connect()
    try:
        if not conn.execute("SELECT id FROM paint_batch WHERE id=?", (bid,)).fetchone():
            return {"ok": False, "error": "배치를 찾을 수 없습니다."}
        if "note" in data:
            conn.execute("UPDATE paint_batch SET good_qty=?,defect_qty=?,note=? WHERE id=?",
                         (good_qty, dq, (data.get("note") or "").strip(), bid))
        else:
            conn.execute("UPDATE paint_batch SET good_qty=?,defect_qty=? WHERE id=?", (good_qty, dq, bid))
        conn.execute("DELETE FROM paint_batch_defect WHERE batch_id=?", (bid,))
        for iid, q in defects:
            conn.execute("INSERT INTO paint_batch_defect(batch_id,inspection_item_id,defect_qty) VALUES (?,?,?)",
                         (bid, iid, q))
        conn.commit()
        return {"ok": True, "good_qty": good_qty, "defect_qty": dq}
    finally:
        conn.close()


def delete(data: dict) -> dict:
    bid = data.get("id")
    conn = db.connect()
    try:
        conn.execute("DELETE FROM paint_batch_defect WHERE batch_id=?", (bid,))
        cur = conn.execute("DELETE FROM paint_batch WHERE id=?", (bid,))
        conn.commit()
        if cur.rowcount == 0:
            return {"ok": False, "error": "배치를 찾을 수 없습니다."}
        return {"ok": True}
    finally:
        conn.close()


# ── 우경 엑셀 파싱 ───────────────────────────────────────────
def _load_wb(b64: str):
    try:
        import openpyxl
    except ImportError:
        return None, "openpyxl 라이브러리가 없습니다."
    try:
        raw = base64.b64decode((b64 or "").split(",")[-1])
        return openpyxl.load_workbook(io.BytesIO(raw), data_only=True), None
    except Exception as exc:  # noqa: BLE001
        return None, f"엑셀을 읽지 못했습니다: {exc}"


def _parse_sheet(ws) -> dict | None:
    """반납일별 시트 파싱 → {good, defects:[{name,qty}], total}. 못 읽으면 None."""
    sr = None
    for r in range(1, min(ws.max_row, 40) + 1):
        b = ws.cell(r, 2).value
        if b and "소" in str(b) and "계" in str(b):   # '소   계'
            sr = r
            break
    if not sr:
        return None
    hr = sr - 1
    good = rules._to_int(ws.cell(sr, 3).value)   # C열=양품
    defects = []
    for col in range(4, ws.max_column + 1):       # D열~ = 불량 유형
        name = ws.cell(hr, col).value
        qty = rules._to_int(ws.cell(sr, col).value)
        if name and str(name).strip() and qty > 0:
            defects.append({"name": str(name).strip(), "qty": qty})
    total = good + sum(d["qty"] for d in defects)
    return {"good": good, "defects": defects, "total": total}


def preview_excel(data: dict) -> dict:
    wb, err = _load_wb(data.get("content"))
    if err:
        return {"ok": False, "error": err}
    prism_id = data.get("prism_id")
    conn = db.connect()
    try:
        existing = set()
        if prism_id:
            for r in conn.execute("SELECT batch_date FROM paint_batch WHERE prism_id=?", (prism_id,)).fetchall():
                existing.add(str(r["batch_date"]))
        batches = []
        for name in wb.sheetnames:
            if name.strip().lower() == "summary":
                continue
            parsed = _parse_sheet(wb[name])
            if not parsed:
                continue
            batches.append({"batch_date": name.strip(), **parsed,
                            "already": name.strip() in existing})
    finally:
        conn.close()
    tot_good = sum(b["good"] for b in batches if not b["already"])
    tot_def = sum(d["qty"] for b in batches if not b["already"] for d in b["defects"])
    return {"ok": True, "batches": batches,
            "summary": {"new_batches": len([b for b in batches if not b["already"]]),
                        "skip": len([b for b in batches if b["already"]]),
                        "good": tot_good, "defect": tot_def}}


def commit_excel(data: dict) -> dict:
    wb, err = _load_wb(data.get("content"))
    if err:
        return {"ok": False, "error": err}
    prism_id = data.get("prism_id")
    if not prism_id:
        return {"ok": False, "error": "도장완료 프리즘을 선택하세요."}
    conn = db.connect()
    try:
        ts = db.now_str()
        existing = {str(r["batch_date"]) for r in
                    conn.execute("SELECT batch_date FROM paint_batch WHERE prism_id=?", (prism_id,)).fetchall()}
        made = 0
        skipped = 0
        for name in wb.sheetnames:
            if name.strip().lower() == "summary":
                continue
            bd = name.strip()
            if bd in existing:
                skipped += 1
                continue
            parsed = _parse_sheet(wb[name])
            if not parsed:
                continue
            dq = sum(d["qty"] for d in parsed["defects"])
            cur = conn.execute(
                "INSERT INTO paint_batch(prism_id,batch_date,good_qty,defect_qty,note,created_at) VALUES (?,?,?,?,?,?)",
                (prism_id, bd, parsed["good"], dq, "우경 불량자료 업로드", ts))
            bid = cur.lastrowid
            for d in parsed["defects"]:
                iid = _find_or_create_item(conn, d["name"], ts)
                conn.execute("INSERT INTO paint_batch_defect(batch_id,inspection_item_id,defect_qty) VALUES (?,?,?)",
                             (bid, iid, d["qty"]))
            made += 1
        conn.commit()
        return {"ok": True, "created": made, "skipped": skipped}
    finally:
        conn.close()
