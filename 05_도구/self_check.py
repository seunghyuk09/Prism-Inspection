# -*- coding: utf-8 -*-
"""자동 통합 점검 — 코드를 고친 뒤 '다 잘 돌아가나?' 한 번에 확인.

점검 내용:
  1) 모든 백엔드 모듈 import (문법/임포트 오류 탐지)
  2) 규칙(rules) 술어 단위 점검 (참/거짓이 의도대로인지)
  3) 임시 DB 로 기준정보 CRUD 동작 (운영 DB 02_DB/prism.sqlite 는 절대 건드리지 않음)

사용:  python self_check.py
끝에 [PASS]/[FAIL] 요약. 실패가 1개라도 있으면 종료코드 1.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

# backend 모듈 경로 추가
BACKEND = Path(__file__).resolve().parents[1] / "04_웹앱" / "backend"
sys.path.insert(0, str(BACKEND))

PASS, FAIL = 0, 0
FAILS: list[str] = []


def check(name: str, cond: bool):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  [OK]   {name}")
    else:
        FAIL += 1
        FAILS.append(name)
        print(f"  [FAIL] {name}")


def main():
    print("== 1) 모듈 import ==")
    import db
    import rules
    import supplier_service
    import prism_service
    import inspection_item_service
    import product_service
    import receipt_service
    import inspection_service
    import paint_service
    import stock_service
    import plan_service
    import report_service
    import auth
    import local_server  # 라우트 등록까지 로드되는지
    check("backend 모듈 전체 import", True)
    check("POST 라우트 등록됨", len(local_server.POST_ROUTES) >= 8)

    print("== 2) 규칙(rules) 술어 ==")
    check("can_create_new_receipt(ACTIVE)=참", rules.can_create_new_receipt("ACTIVE") is True)
    check("can_create_new_receipt(REMNANT)=거짓", rules.can_create_new_receipt("REMNANT") is False)
    check("is_plastic(PLASTIC)=참", rules.is_plastic("PLASTIC") is True)
    check("product_uses_prism(None)=거짓", rules.product_uses_prism(None) is False)
    check("plan_consumption: 100×2(프리즘있음)=200", rules.plan_consumption_qty(100, 2, 5) == 200)
    check("plan_consumption: 미사용제품=0", rules.plan_consumption_qty(100, 2, None) == 0)
    # 수량 정합성
    check("qty_balanced 완료 1960+40==2000", rules.qty_balanced(1960, 40, 2000, True) is True)
    check("qty_balanced 미완료 1930+20<=1950(soft)", rules.qty_balanced(1930, 20, 1950, False) is True)
    check("qty_balanced 완료 불일치=거짓", rules.qty_balanced(1930, 30, 1950, True) is False)  # 1960!=1950
    check("paint_send_valid 1960<=1960", rules.paint_send_valid(1960, 1960) is True)
    check("paint_send_valid 2000>1960=거짓", rules.paint_send_valid(2000, 1960) is False)
    check("can_create_paint_job(PLASTIC,완료)=참", rules.can_create_paint_job("PLASTIC", True) is True)
    check("can_create_paint_job(GLASS,완료)=거짓", rules.can_create_paint_job("GLASS", True) is False)
    v, _ = rules.validate_inspection_item("테스트", 0, 0)
    check("검사항목 둘다 미적용=검증실패", v is False)

    print("== 3) 임시 DB 로 기준정보 CRUD ==")
    tmp = Path(tempfile.mkdtemp(prefix="prism_check_"))
    db.DB_DIR = tmp                      # 운영 DB 대신 임시 경로로 교체
    db.DB_PATH = tmp / "check.sqlite"
    db.init_db()
    check("임시 DB 시드: 공급사 4", len(supplier_service.list_all()) == 4)
    check("임시 DB 시드: 검사항목 5", len(inspection_item_service.list_all()) == 5)
    check("임시 DB 시드: 프리즘 8", len(prism_service.list_all()) == 8)
    by_code = {p["item_code"]: p for p in prism_service.list_all()}
    check("페어: 52791 → SP10-2103-ALT", by_code["52791"]["painted_into_code"] == "SP10-2103-ALT")
    check("페어: A0001 → SP10-2103", by_code["A0001"]["painted_into_code"] == "SP10-2103")
    check("52791 도장상태 RAW + 공급사 ALT", by_code["52791"]["paint_state"] == "RAW" and by_code["52791"]["supplier_name"] == "ALT")
    check("00015(U30) 유리 + 도장 NONE + TianCheng", by_code["00015"]["prism_type"] == "GLASS" and by_code["00015"]["paint_state"] == "NONE" and by_code["00015"]["supplier_name"] == "TianCheng")
    check("도장완료 코드는 입고불가 규칙", rules.is_receivable_prism("PAINTED") is False)

    # 공급사 등록 + 활성 공급사 신규입고 규칙 연계 확인
    r = supplier_service.create({"name": "테스트공급사", "status": "ACTIVE", "note": ""})
    check("공급사 등록 성공", r.get("ok") is True)
    check("공급사 중복 등록 차단", supplier_service.create({"name": "테스트공급사", "status": "ACTIVE"}).get("ok") is False)

    # 검사항목 추가(확장형) → 6개
    r = inspection_item_service.create({"name": "이물질", "category": "외관",
                                        "applies_to_incoming": 1, "applies_to_post_paint": 1})
    check("검사항목 추가 성공", r.get("ok") is True)
    check("검사항목 6개로 증가", len(inspection_item_service.list_all()) == 6)

    # 제품마스터: 프리즘 사용 제품 + 미사용 제품 (RAW 플라스틱 52791 사용)
    plastic = by_code["52791"]
    r = product_service.create({"product_code": "U20", "product_name": "U20",
                                "prism_id": plastic["id"], "prism_per_unit": 1})
    check("제품(프리즘사용) 등록 성공", r.get("ok") is True)
    r = product_service.create({"product_code": "ID USB SC", "product_name": "ID USB SC",
                                "prism_id": "", "prism_per_unit": 0})
    check("제품(프리즘미사용) 등록 성공", r.get("ok") is True)
    prod_u20 = next((p for p in product_service.list_all() if p["product_code"] == "U20"), None)
    check("제품 U20 의 프리즘 연결됨", prod_u20 and prod_u20["prism_id"] == plastic["id"])
    # 회귀: 프리즘 사용 제품에 소요량 0(명시) → 검증 실패해야 함(0이 1로 덮어써지지 않게)
    bad = product_service.create({"product_code": "__BAD__", "prism_id": plastic["id"], "prism_per_unit": 0})
    check("프리즘사용 + 소요량0(명시) → 검증실패", bad.get("ok") is False)

    # 사용여부 토글
    r = product_service.set_active({"id": prod_u20["id"], "is_active": 0})
    check("제품 사용여부 토글 성공", r.get("ok") is True)

    print("== 4) 임시 DB 로 입고 / 로트 ==")
    alt = next((s for s in supplier_service.list_all() if s["name"] == "ALT"), None)
    ofren = next((s for s in supplier_service.list_all() if s["name"] == "오프렌"), None)
    rc = receipt_service.create({"receipt_date": "2026-06-20", "prism_id": plastic["id"],
                                 "supplier_id": alt["id"], "received_qty": 5000})
    check("입고 등록(ALT) 성공", rc.get("ok") is True)
    rid = rc.get("receipt_id")
    bad_rc = receipt_service.create({"receipt_date": "2026-06-20", "prism_id": plastic["id"],
                                     "supplier_id": ofren["id"], "received_qty": 100})
    check("오프렌(REMNANT) 신규입고 차단", bad_rc.get("ok") is False)
    check("로트1 추가(2000)", receipt_service.add_lot({"receipt_id": rid, "lot_qty": 2000}).get("ok") is True)
    check("로트2 추가(2000)", receipt_service.add_lot({"receipt_id": rid, "lot_qty": 2000}).get("ok") is True)
    check("로트 초과분리 차단(잔여 1000)", receipt_service.add_lot({"receipt_id": rid, "lot_qty": 2000}).get("ok") is False)
    check("로트3 추가(잔여 1000 소진)", receipt_service.add_lot({"receipt_id": rid, "lot_qty": 1000}).get("ok") is True)
    det = receipt_service.detail(rid)
    check("분리합계 5000 / 잔여 0", det["lot_qty_sum"] == 5000 and det["remaining_qty"] == 0)
    # 도장완료 코드는 직접 입고 불가
    check("도장완료 코드 입고 차단",
          receipt_service.create({"receipt_date": "2026-06-20", "prism_id": by_code["SP10-2103-ALT"]["id"], "received_qty": 100}).get("ok") is False)
    # 공급사 미지정 → 프리즘(52791)에서 ALT 자동 결정
    check("공급사 자동결정(52791→ALT) 입고",
          receipt_service.create({"receipt_date": "2026-06-21", "prism_id": by_code["52791"]["id"], "received_qty": 1000}).get("ok") is True)

    print("== 5) 임시 DB 로 검사 / 페인트 ==")
    lid = receipt_service.detail(rid)["lots"][0]["id"]   # 첫 로트(2000)
    inc_item = next(it["id"] for it in inspection_item_service.list_all() if it["applies_to_incoming"])
    r = inspection_service.save({"lot_id": lid, "stage": "INCOMING", "is_complete": 1,
                                 "defects": [{"item_id": inc_item, "qty": 40}]})
    check("입고검사 저장(완료) 양품=1960", r.get("ok") and r.get("good_qty") == 1960)
    check("페인트 발송 양품초과 차단", paint_service.create_job({"lot_id": lid, "sent_date": "2026-06-15", "sent_qty": 2000}).get("ok") is False)
    check("페인트 발송 1960", paint_service.create_job({"lot_id": lid, "sent_date": "2026-06-15", "sent_qty": 1960}).get("ok") is True)
    job_id = inspection_service.get_for_lot({"id": lid})["paint"]["jobs"][0]["id"]
    check("페인트 회수 1000(부분)", paint_service.add_return({"paint_job_id": job_id, "returned_date": "2026-06-18", "returned_qty": 1000}).get("ok") is True)
    check("페인트 회수 초과 차단", paint_service.add_return({"paint_job_id": job_id, "returned_date": "2026-06-19", "returned_qty": 1000}).get("ok") is False)
    check("페인트 회수 960(완료)", paint_service.add_return({"paint_job_id": job_id, "returned_date": "2026-06-20", "returned_qty": 960}).get("ok") is True)
    d2 = inspection_service.get_for_lot({"id": lid})
    check("회수완료 is_returned=1", d2["paint"]["jobs"][0]["is_returned"] == 1)
    check("페인트후검사 base=회수합 1960", d2["post_paint"]["base_qty"] == 1960)
    post_item = next(it["id"] for it in inspection_item_service.list_all() if it["applies_to_post_paint"])
    check("페인트후검사 미완료(점진) 저장", inspection_service.save({"lot_id": lid, "stage": "POST_PAINT", "is_complete": 0, "defects": [{"item_id": post_item, "qty": 10}]}).get("ok") is True)
    rp = inspection_service.save({"lot_id": lid, "stage": "POST_PAINT", "is_complete": 1, "defects": [{"item_id": post_item, "qty": 30}]})
    check("페인트후검사 완료 양품=1930", rp.get("ok") and rp.get("good_qty") == 1930)
    check("최종양품=1930", inspection_service.get_for_lot({"id": lid})["final_good"] == 1930)
    # 글래스 로트는 페인트후검사 차단
    glass = next(p for p in prism_service.list_all() if p["prism_type"] == "GLASS")
    gr = receipt_service.create({"receipt_date": "2026-06-20", "prism_id": glass["id"], "supplier_id": alt["id"], "received_qty": 100})
    receipt_service.add_lot({"receipt_id": gr["receipt_id"], "lot_qty": 100})
    glid = receipt_service.detail(gr["receipt_id"])["lots"][0]["id"]
    check("글래스 페인트후검사 차단", inspection_service.save({"lot_id": glid, "stage": "POST_PAINT", "is_complete": 0, "defects": []}).get("ok") is False)

    print("== 6) 임시 DB 로 잔량 / 구매계획 ==")
    def stock_by_code():
        return {r["item_code"]: r for r in stock_service.stock_status()["items"]}
    st = stock_by_code()
    # 52791(RAW): 입고양품 1960 − 페인트발송 1960 = 0
    check("52791(RAW) 현재고 0", st["52791"]["on_hand"] == 0)
    # SP10-2103-ALT(PAINTED): 페인트후 양품 1930
    check("SP10-2103-ALT(도장완료) 현재고 1930", st["SP10-2103-ALT"]["on_hand"] == 1930)
    # 기초재고: 오프렌 도장완료 SP10-2103 에 1200
    stock_service.add_adjustment({"prism_id": by_code["SP10-2103"]["id"], "qty": 1200, "reason": "OPENING"})
    check("기초재고 SP10-2103 = 1200", stock_by_code()["SP10-2103"]["on_hand"] == 1200)

    # 구매계획 임포트: 도장완료 코드를 소비하는 제품 등록 후 임시 엑셀 임포트
    product_service.create({"product_code": "TESTPROD", "prism_id": by_code["SP10-2103-ALT"]["id"], "prism_per_unit": 1})
    import openpyxl, io as _io, base64 as _b64
    _wb = openpyxl.Workbook(); _ws = _wb.active; _ws.title = "PlanTest"
    _ws.cell(row=7, column=3, value="TESTPROD"); _ws.cell(row=7, column=5, value=500)   # C7/E7
    _ws.cell(row=8, column=3, value="UNKNOWNPROD"); _ws.cell(row=8, column=5, value=300)  # 미등록
    _buf = _io.BytesIO(); _wb.save(_buf); _b = _b64.b64encode(_buf.getvalue()).decode()
    pv = plan_service.preview({"content": _b, "sheet": "PlanTest"})
    check("플랜 미리보기: 매칭1 · 소비500", pv["summary"]["matched"] == 1 and pv["summary"]["total_consumed"] == 500)
    check("플랜 미리보기: 미등록 1건 감지", pv["summary"]["unmatched"] == 1)
    cm = plan_service.commit({"content": _b, "sheet": "PlanTest", "plan_month": "2026-06"})
    check("플랜 확정: 소비 500", cm.get("ok") and cm.get("consumed_total") == 500)
    check("소비 후 SP10-2103-ALT = 1430", stock_by_code()["SP10-2103-ALT"]["on_hand"] == 1430)
    plan_service.delete_plan({"id": cm["plan_id"]})
    check("계획 취소 후 복구 1930", stock_by_code()["SP10-2103-ALT"]["on_hand"] == 1930)

    print("== 7) 집계 / 엑셀 ==")
    rep = report_service.summary()
    check("집계 by_lot 1건 이상", len(rep["by_lot"]) >= 1)
    lot_row = next((b for b in rep["by_lot"] if b["lot_id"] == lid), None)
    check("집계 로트 최종양품 1930 / 수율 96.5", lot_row and lot_row["final_good"] == 1930 and lot_row["yield"] == 96.5)
    item_row = next((b for b in rep["by_item"] if b["name"] == "원자재"), None)
    check("집계 원자재 입고불량 40", item_row and item_row["inc_defect"] == 40)
    sup_row = next((b for b in rep["by_supplier"] if b["supplier_name"] == "ALT"), None)
    check("집계 공급사 ALT 존재", sup_row is not None)
    xb, xfn = report_service.lot_excel(lid)
    check("로트 검사이력 엑셀 생성(xlsx)", bool(xb) and xb[:2] == b"PK" and xfn.endswith(".xlsx"))
    rb, rfn = report_service.report_excel()
    check("전체 집계 엑셀 생성(xlsx)", bool(rb) and rb[:2] == b"PK")

    print("== 8) 인증(로그인/세션) ==")
    import tempfile as _tf
    udir = Path(_tf.mkdtemp(prefix="prism_auth_"))
    us = auth.UserStore(udir / "users.json")
    check("기본 관리자 시드 + 로그인", us.verify("admin", "secugen") is True)
    check("틀린 비번 거부", us.verify("admin", "wrong") is False)
    check("admin 은 관리자 역할", us.is_admin("admin") is True)
    check("사용자 추가", us.add("hong", "홍길동", "pw1", "user")[0] is True)
    check("추가 사용자 로그인", us.verify("hong", "pw1") is True)
    check("추가 사용자는 일반(user)", us.is_admin("hong") is False)
    # 3단계 역할 + 권한
    check("admin 은 담당자이상(쓰기가능)", us.is_manager_or_above("admin") is True)
    check("user 는 담당자미만(조회전용)", us.is_manager_or_above("hong") is False)
    check("담당자(manager) 추가", us.add("kim", "김담당", "pw2", "manager")[0] is True)
    check("담당자는 쓰기가능", us.is_manager_or_above("kim") is True)
    check("담당자는 사용자관리 불가(관리자 아님)", us.is_admin("kim") is False)
    check("관리자가 역할 수정: hong→담당자", us.update_user("hong", role="manager")[0] is True and us.is_manager_or_above("hong") is True)
    check("관리자가 이름 수정", us.update_user("hong", name="홍길동2")[0] is True and us.get("hong")["name"] == "홍길동2")
    check("마지막 관리자 강등 차단", us.update_user("admin", role="manager")[0] is False)
    check("중복 아이디 차단", us.add("hong", "x", "y")[0] is False)
    check("마지막 관리자 삭제 차단", us.remove("admin")[0] is False)
    check("비번 변경", us.change_password("hong", "newpw")[0] is True and us.verify("hong", "newpw") is True)
    check("ID 변경 + 재해시 로그인", us.change_id("hong", "hong2", "newpw")[0] is True and us.verify("hong2", "newpw") is True and us.get("hong") is None)
    check("ID 변경 틀린 비번 거부", us.change_id("hong2", "hong3", "wrong")[0] is False)
    sm = auth.SessionManager()
    tok = sm.issue("hong")
    check("세션 발급/검증", sm.validate(tok) == "hong")
    check("없는 토큰은 None", sm.validate("nope") is None)
    sm.revoke(tok)
    check("세션 폐기 후 무효", sm.validate(tok) is None)
    import shutil as _sh
    _sh.rmtree(udir, ignore_errors=True)

    # 정리
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)

    print("\n========================================")
    print(f"  결과: PASS {PASS} / FAIL {FAIL}")
    if FAILS:
        print("  실패 항목:")
        for f in FAILS:
            print("   -", f)
    print("========================================")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
